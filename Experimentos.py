from clases.AlgorithmsML.DecisionTree import *
from clases.AlgorithmsML.GaussianNaiveBayes import *
from clases.AlgorithmsML.LogisticRegresion import *
from clases.AlgorithmsML.MultiLayerPerceptron import *
from clases.AlgorithmsML.RandomForest import *
from clases.AlgorithmsML.XGBoost import *
from clases.AlgorithmsML.KNearestNeighbor import *
from clases.AlgorithmsDL.ConvolutionNeuralNetworkClass import *
from clases.AlgorithmsDL.AutoencoderClass import *
from clases.AlgorithmsDL.DenoisingAutoencoderClass import *
from clases.AlgorithmsDL.RecurrentNeuralNetworkClass import *
from clases.AlgorithmsDL.LongShortTermMemoryClass import *
from clases.AlgorithmsDL.BackpropagationNeuralNetworkClass import *
from clases.ImbalancedPerformance import ImbalancedPerformanceClass
from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np


def plot_bar_test(ip, names, title, pos, path):
    x = np.arange(len(names))
    width = 0.15
    accuracy_1 = []
    recall_1 = []
    precision_1 = []
    f1_1 = []
    fig, ax = plt.subplots()
    for j in pos:
        accuracy_1.append(ip.accuracy_tests[j])
        precision_1.append(ip.precision_tests[j])
        recall_1.append(ip.recall_tests[j])
        f1_1.append(ip.f1_score_tests[j])
    ax.bar(x - width, accuracy_1, width, label='Accuracy Test')
    ax.bar(x - width/2, precision_1, width, label='Precision Test')
    ax.bar(x + width/2, recall_1, width, label='Recall Test')
    ax.bar(x + width, f1_1, width, label='F1 Score Test')
    ax.set_ylabel('Scores')
    ax.set_title(title)
    ax.set_xticks(x)
    ax.set_xticklabels(names)
    ax.legend()
    fig.tight_layout()
    plt.ylim((0.99, 1.0))
    plt.savefig(path)
    plt.show()


def plot_bar_test_full(ip, names, title, pos, path):
    x = np.arange(len(names))
    width = 0.15
    accuracy_1 = []
    recall_1 = []
    precision_1 = []
    f1_1 = []
    fig, ax = plt.subplots()
    for j in pos:
        accuracy_1.append(ip.accuracy_tests[j])
        precision_1.append(ip.precision_tests[j])
        recall_1.append(ip.recall_tests[j])
        f1_1.append(ip.f1_score_tests[j])
    ax.bar(x - width, accuracy_1, width, label='Accuracy Test')
    ax.bar(x - width/2, precision_1, width, label='Precision Test')
    ax.bar(x + width/2, recall_1, width, label='Recall Test')
    ax.bar(x + width, f1_1, width, label='F1 Score Test')
    ax.set_ylabel('Scores')
    ax.set_title(title)
    ax.set_xticks(x)
    ax.set_xticklabels(names)
    ax.legend()
    fig.tight_layout()
    plt.savefig(path)
    plt.show()


def plot_bar_train(ip, names, title, pos, path):
    x = np.arange(len(names))
    width = 0.15
    accuracy_1 = []
    recall_1 = []
    precision_1 = []
    f1_1 = []
    fig, ax = plt.subplots()
    for j in pos:
        accuracy_1.append(ip.accuracy_train[j])
        precision_1.append(ip.precision_train[j])
        recall_1.append(ip.recall_train[j])
        f1_1.append(ip.f1_score_train[j])
    ax.bar(x - width, accuracy_1, width, label='Accuracy Train')
    ax.bar(x - width/2, precision_1, width, label='Precision Train')
    ax.bar(x + width/2, recall_1, width, label='Recall Train')
    ax.bar(x + width, f1_1, width, label='F1 Score Train')
    ax.set_ylabel('Scores')
    ax.set_title(title)
    ax.set_xticks(x)
    ax.set_xticklabels(names)
    ax.legend()
    fig.tight_layout()
    plt.savefig(path)
    plt.show()


def plot_graph(ip, names, title, pos, epochs, path):
    plt.ylabel('f1 score')
    plt.xlabel('epochs')
    legend = []
    epoch_range = range(0, epochs)
    i = 0
    for j in pos:
        plt.plot(epoch_range, ip.histories[j].history['custom_f1'])
        plt.plot(epoch_range, ip.histories[j].history['val_custom_f1'])
        legend.append(names[i] + ' Train')
        legend.append(names[i] + ' Test')
        i = i + 1
    plt.legend(legend, loc='upper left')
    plt.title(title)
    plt.savefig(path)
    plt.show()


def experimento1(path_save="PNG/Experimento1/KNN_Test_size.png"):
    ip = ImbalancedPerformanceClass()
    ip.solve_imbalanced("creditcard.csv", 0.1)
    ip.performanceML(KNNImbalanced(ip))
    ip.solve_imbalanced("creditcard.csv", 0.2)
    ip.performanceML(KNNImbalanced(ip))
    ip.solve_imbalanced("creditcard.csv", 0.3)
    ip.performanceML(KNNImbalanced(ip))
    ip.solve_imbalanced("creditcard.csv", 0.4)
    ip.performanceML(KNNImbalanced(ip))
    ip.solve_imbalanced("creditcard.csv", 0.5)
    ip.performanceML(KNNImbalanced(ip))

    x = np.arange(len(ip.test_size))
    width = 0.15
    fig, ax = plt.subplots()
    ax.bar(x - width * 2, ip.accuracy_tests, width, label='Accuracy Test')
    ax.bar(x - width, ip.aucs_tests, width, label='AUC Test')
    ax.bar(x, ip.precision_tests, width, label='Precision Test')
    ax.bar(x + width, ip.recall_tests, width, label='Recall Test')
    ax.bar(x + width * 2, ip.f1_score_tests, width, label='F1 Score Test')
    ax.set_ylabel('Scores')
    ax.set_title('Scores by test size in KNN Imbalanced')
    ax.set_xticks(x)
    ax.set_xticklabels(ip.test_size)
    ax.legend()

    fig.tight_layout()
    plt.savefig(path_save)
    plt.show()

    ip.show_comparison_test_size()


def experimento2(path_save="PNG/Experimento2/"):
    ip = ImbalancedPerformanceClass()
    ip.solve_imbalanced("creditcard.csv")
    # Imbalanced
    ip.performanceML(DTImbalanced(ip))
    ip.performanceML(GNBImbalanced(ip))
    ip.performanceML(KNNImbalanced(ip))
    ip.performanceML(LRImbalanced(ip))
    ip.performanceML(MLPImbalanced(ip))
    ip.performanceML(RFImbalanced(ip))
    ip.performanceML(XGBImbalanced(ip))
    # UnderSampler
    ip.performanceML(DTUnderSample(ip))
    ip.performanceML(GNBUnderSample(ip))
    ip.performanceML(KNNUnderSample(ip))
    ip.performanceML(LRUnderSample(ip))
    ip.performanceML(MLPUnderSample(ip))
    ip.performanceML(RFUnderSample(ip))
    ip.performanceML(XGBUnderSample(ip))
    # OverSampler
    ip.performanceML(DTOverSample(ip))
    ip.performanceML(GNBOverSample(ip))
    ip.performanceML(KNNOverSample(ip))
    ip.performanceML(LROverSample(ip))
    ip.performanceML(MLPOverSample(ip))
    ip.performanceML(RFOverSample(ip))
    ip.performanceML(XGBOverSample(ip))
    # SMOTE
    ip.performanceML(DTSMOTE(ip))
    ip.performanceML(GNBSMOTE(ip))
    ip.performanceML(KNNSMOTE(ip))
    ip.performanceML(LRSMOTE(ip))
    ip.performanceML(MLPSMOTE(ip))
    ip.performanceML(RFSMOTE(ip))
    ip.performanceML(XGBSMOTE(ip))
    # ADASYN
    ip.performanceML(DTADASYN(ip))
    ip.performanceML(GNBADASYN(ip))
    ip.performanceML(KNNADASYN(ip))
    ip.performanceML(LRADASYN(ip))
    ip.performanceML(MLPADASYN(ip))
    ip.performanceML(RFADASYN(ip))
    ip.performanceML(XGBADASYN(ip))

    comparision = ip.show_comparison()

    names_1 = ["DT", "GNB", "KNN", "LR", "MLP", "RF", "XGBOOST"]
    names_2 = ["Imbalanced", "UnderSampler", "OverSampler", "SMOTE", "ADASYN"]

    # Strategies by models
    # Imbalanced
    title = "Scores by metrics in ML models with Imbalanced"
    pos = [0, 1, 2, 3, 4, 5, 6]
    path = path_save + "ML_imbalanced_test"
    plot_bar_test_full(ip, names_1, title, pos, path)
    # UnderSampler
    title = "Scores by metrics in ML models with UnderSampler"
    pos = [7, 8, 9, 10, 11, 12, 13]
    path = path_save + "ML_UnderSampler_test"
    plot_bar_test_full(ip, names_1, title, pos, path)
    # OverSampler
    title = "Scores by metrics in ML models with OverSampler"
    pos = [14, 15, 16, 17, 18, 19, 20]
    path = path_save + "ML_OverSampler_test"
    plot_bar_test_full(ip, names_1, title, pos, path)
    # SMOTE
    title = "Scores by metrics in ML models with SMOTE"
    pos = [21, 22, 23, 24, 25, 26, 27]
    path = path_save + "ML_SMOTE_test"
    plot_bar_test_full(ip, names_1, title, pos, path)
    # ADASYN
    title = "Scores by metrics in ML models with SMOTE"
    pos = [28, 29, 30, 31, 32, 33, 34]
    path = path_save + "ML_ADASYN_test"
    plot_bar_test_full(ip, names_1, title, pos, path)


    book = load_workbook('Experimentos.xlsx')
    comparision = comparision.sort_values('F1 Score', ascending=False)
    writer = pd.ExcelWriter('Experimentos.xlsx', engine='openpyxl', if_sheet_exists='replace')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    comparision.to_excel(writer, sheet_name='Experimento2', index=False)
    writer.save()
    writer.close()


def experimento3(path_save="PNG/Experimento3/"):
    ip = ImbalancedPerformanceClass()
    ip.solve_imbalanced("creditcard.csv")
    ip.epochs = 1
    ip.performanceCNN(CNNImbalanced(ip), [0.5, 0.4, 0.3])
    ip.performanceRNN(RNNImbalanced(ip), [0.5, 0.4, 0.3])
    ip.performanceCNN(CNNImbalanced(ip), [0.3, 0.4, 0.5])
    ip.performanceRNN(RNNImbalanced(ip), [0.3, 0.4, 0.5])
    ip.performanceCNN(CNNImbalanced(ip), [0.5, 0.5, 0.5])
    ip.performanceRNN(RNNImbalanced(ip), [0.5, 0.5, 0.5])
    ip.performanceCNN(CNNImbalanced(ip), [0.4, 0.4, 0.4])
    ip.performanceRNN(RNNImbalanced(ip), [0.4, 0.4, 0.4])
    ip.performanceCNN(CNNImbalanced(ip), [0.3, 0.3, 0.3])
    ip.performanceRNN(RNNImbalanced(ip), [0.3, 0.3, 0.3])
    ip.performanceCNN(CNNImbalanced(ip), [0.2, 0.2, 0.2])
    ip.performanceRNN(RNNImbalanced(ip), [0.2, 0.2, 0.2])
    ip.performanceCNN(CNNImbalanced(ip), [0.2, 0.2, 0.5])
    ip.performanceRNN(RNNImbalanced(ip), [0.2, 0.2, 0.5])
    ip.performanceCNN(CNNImbalanced(ip), [0.5, 0.2, 0.2])
    ip.performanceRNN(RNNImbalanced(ip), [0.5, 0.2, 0.2])

    p = [0, 2, 4, 6, 8, 10, 12, 14]
    ip = [1, 3, 5, 7, 9, 11, 13, 15]
    x = np.arange(8)
    names1 = []
    accuracy1 = []
    f11 =[]
    names2 = []
    accuracy2 = []
    f12 = []
    width = 0.15
    for i in p:
        if i % 2 == 0:
            names1.append(ip.dropouts[i])
        else:
            names2.append(ip.dropouts[i])
    fig, ax = plt.subplots()
    for i in p:
        if i % 2 == 0:
            accuracy1.append(ip.accuracy_tests[i])
            f11.append(ip.f1_score_tests[i])
    ax.bar(x - width, accuracy1, width, label='Accuracy Test')
    ax.bar(x + width, f11, width, label='F1 Score Test')
    ax.set_ylabel('Scores')
    ax.set_title('Scores by metrics in CNN Imbalanced with dropouts')
    ax.set_xticks(x)
    ax.set_xticklabels(names1)
    ax.legend()
    fig.tight_layout()
    plt.ylim((0.995, 1.0))
    plt.savefig(path_save + "CNN_dropouts_tests")
    plt.show()
    fig, ax = plt.subplots()
    for i in p:
        if i % 2 != 0:
            accuracy2.append(ip.accuracy_tests[i])
            f12.append(ip.f1_score_tests[i])
    ax.bar(x - width, accuracy2, width, label='Accuracy Test')
    ax.bar(x + width, f12, width, label='F1 Score Test')
    ax.set_ylabel('Scores')
    ax.set_title('Scores by metrics in RNN Imbalanced with dropouts')
    ax.set_xticks(x)
    ax.set_xticklabels(names2)
    ax.legend()
    fig.tight_layout()
    plt.ylim((0.995, 1.0))
    plt.savefig(path_save + "RNN_dropouts_tests")
    plt.show()

    comparision = ip.show_comparison_dropout()
    values = comparision.values
    df1 = pd.DataFrame(values[:2], columns=['Dropout', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                            'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df2 = pd.DataFrame(values[2:4], columns=['Dropout', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                             'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df3 = pd.DataFrame(values[4:6], columns=['Dropout', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                             'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df4 = pd.DataFrame(values[6:8], columns=['Dropout', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                             'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df5 = pd.DataFrame(values[8:10], columns=['Dropout', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                              'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df6 = pd.DataFrame(values[10:12], columns=['Dropout', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                               'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df7 = pd.DataFrame(values[12:14], columns=['Dropout', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                               'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df8 = pd.DataFrame(values[14:16], columns=['Dropout', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                               'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    comparision = comparision.sort_values('F1 Score', ascending=False)
    book = load_workbook('Experimentos.xlsx')
    writer = pd.ExcelWriter('Experimentos.xlsx', engine='openpyxl', if_sheet_exists='replace', mode='a')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df1.to_excel(writer, sheet_name='Experimento3', index=False, startrow=0)
    df2.to_excel(writer, sheet_name='Experimento3', index=False, startrow=4)
    df3.to_excel(writer, sheet_name='Experimento3', index=False, startrow=8)
    df4.to_excel(writer, sheet_name='Experimento3', index=False, startrow=12)
    df5.to_excel(writer, sheet_name='Experimento3', index=False, startrow=16)
    df6.to_excel(writer, sheet_name='Experimento3', index=False, startrow=20)
    df7.to_excel(writer, sheet_name='Experimento3', index=False, startrow=24)
    df8.to_excel(writer, sheet_name='Experimento3', index=False, startrow=28)
    comparision.to_excel(writer, sheet_name='Experimento3', index=False, startcol=8)
    writer.save()
    writer.close()


def experimento4_5(path_save="PNG/"):
    ip = ImbalancedPerformanceClass()
    ip.solve_imbalanced("creditcard.csv")
    ip.epochs = 1
    # Imbalanced
    ip.performanceCNN(CNNImbalanced(ip))
    ip.performanceAE(AEImbalanced(ip))
    ip.performanceDAE(DAEImbalanced(ip))
    ip.performanceRNN(RNNImbalanced(ip))
    ip.performanceLSTM(LSTMImbalanced(ip))
    ip.performanceBPNN(BPNNImbalanced(ip))
    # UnderSampler
    ip.performanceCNN(CNNUnderSample(ip))
    ip.performanceAE(AEUnderSample(ip))
    ip.performanceDAE(DAEUnderSample(ip))
    ip.performanceRNN(RNNUnderSample(ip))
    ip.performanceLSTM(LSTMUnderSample(ip))
    ip.performanceBPNN(BPNNUnderSample(ip))
    # OverSampler
    ip.performanceCNN(CNNOverSample(ip))
    ip.performanceAE(AEOverSample(ip))
    ip.performanceDAE(DAEOverSample(ip))
    ip.performanceRNN(RNNOverSample(ip))
    ip.performanceLSTM(LSTMOverSample(ip))
    ip.performanceBPNN(BPNNOverSample(ip))
    # SMOTE
    ip.performanceCNN(CNNSMOTE(ip))
    ip.performanceAE(AESMOTE(ip))
    ip.performanceDAE(DAESMOTE(ip))
    ip.performanceRNN(RNNSMOTE(ip))
    ip.performanceLSTM(LSTMSMOTE(ip))
    ip.performanceBPNN(BPNNSMOTE(ip))
    # ADASYN
    ip.performanceCNN(CNNADASYN(ip))
    ip.performanceAE(AEADASYN(ip))
    ip.performanceDAE(DAEADASYN(ip))
    ip.performanceRNN(RNNADASYN(ip))
    ip.performanceLSTM(LSTMADASYN(ip))
    ip.performanceBPNN(BPNNADASYN(ip))
    ip.epochs = 20
    # Imbalanced
    ip.performanceCNN(CNNImbalanced(ip))
    ip.performanceAE(AEImbalanced(ip))
    ip.performanceDAE(DAEImbalanced(ip))
    ip.performanceRNN(RNNImbalanced(ip))
    ip.performanceLSTM(LSTMImbalanced(ip))
    ip.performanceBPNN(BPNNImbalanced(ip))
    # UnderSampler
    ip.performanceCNN(CNNUnderSample(ip))
    ip.performanceAE(AEUnderSample(ip))
    ip.performanceDAE(DAEUnderSample(ip))
    ip.performanceRNN(RNNUnderSample(ip))
    ip.performanceLSTM(LSTMUnderSample(ip))
    ip.performanceBPNN(BPNNUnderSample(ip))
    # OverSampler
    ip.performanceCNN(CNNOverSample(ip))
    ip.performanceAE(AEOverSample(ip))
    ip.performanceDAE(DAEOverSample(ip))
    ip.performanceRNN(RNNOverSample(ip))
    ip.performanceLSTM(LSTMOverSample(ip))
    ip.performanceBPNN(BPNNOverSample(ip))
    # SMOTE
    ip.performanceCNN(CNNSMOTE(ip))
    ip.performanceAE(AESMOTE(ip))
    ip.performanceDAE(DAESMOTE(ip))
    ip.performanceRNN(RNNSMOTE(ip))
    ip.performanceLSTM(LSTMSMOTE(ip))
    ip.performanceBPNN(BPNNSMOTE(ip))
    # ADASYN
    ip.performanceCNN(CNNADASYN(ip))
    ip.performanceAE(AEADASYN(ip))
    ip.performanceDAE(DAEADASYN(ip))
    ip.performanceRNN(RNNADASYN(ip))
    ip.performanceLSTM(LSTMADASYN(ip))
    ip.performanceBPNN(BPNNADASYN(ip))
    ip.epochs = 50
    # Imbalanced
    ip.performanceCNN(CNNImbalanced(ip))
    ip.performanceAE(AEImbalanced(ip))
    ip.performanceDAE(DAEImbalanced(ip))
    ip.performanceRNN(RNNImbalanced(ip))
    ip.performanceLSTM(LSTMImbalanced(ip))
    ip.performanceBPNN(BPNNImbalanced(ip))
    # UnderSampler
    ip.performanceCNN(CNNUnderSample(ip))
    ip.performanceAE(AEUnderSample(ip))
    ip.performanceDAE(DAEUnderSample(ip))
    ip.performanceRNN(RNNUnderSample(ip))
    ip.performanceLSTM(LSTMUnderSample(ip))
    ip.performanceBPNN(BPNNUnderSample(ip))
    # OverSampler
    ip.performanceCNN(CNNOverSample(ip))
    ip.performanceAE(AEOverSample(ip))
    ip.performanceDAE(DAEOverSample(ip))
    ip.performanceRNN(RNNOverSample(ip))
    ip.performanceLSTM(LSTMOverSample(ip))
    ip.performanceBPNN(BPNNOverSample(ip))
    # SMOTE
    ip.performanceCNN(CNNSMOTE(ip))
    ip.performanceAE(AESMOTE(ip))
    ip.performanceDAE(DAESMOTE(ip))
    ip.performanceRNN(RNNSMOTE(ip))
    ip.performanceLSTM(LSTMSMOTE(ip))
    ip.performanceBPNN(BPNNSMOTE(ip))
    # ADASYN
    ip.performanceCNN(CNNADASYN(ip))
    ip.performanceAE(AEADASYN(ip))
    ip.performanceDAE(DAEADASYN(ip))
    ip.performanceRNN(RNNADASYN(ip))
    ip.performanceLSTM(LSTMADASYN(ip))
    ip.performanceBPNN(BPNNADASYN(ip))
    ip.epochs = 100
    # Imbalanced
    ip.performanceCNN(CNNImbalanced(ip))
    ip.performanceAE(AEImbalanced(ip))
    ip.performanceDAE(DAEImbalanced(ip))
    ip.performanceRNN(RNNImbalanced(ip))
    ip.performanceLSTM(LSTMImbalanced(ip))
    ip.performanceBPNN(BPNNImbalanced(ip))
    # UnderSampler
    ip.performanceCNN(CNNUnderSample(ip))
    ip.performanceAE(AEUnderSample(ip))
    ip.performanceDAE(DAEUnderSample(ip))
    ip.performanceRNN(RNNUnderSample(ip))
    ip.performanceLSTM(LSTMUnderSample(ip))
    ip.performanceBPNN(BPNNUnderSample(ip))
    # OverSampler
    ip.performanceCNN(CNNOverSample(ip))
    ip.performanceAE(AEOverSample(ip))
    ip.performanceDAE(DAEOverSample(ip))
    ip.performanceRNN(RNNOverSample(ip))
    ip.performanceLSTM(LSTMOverSample(ip))
    ip.performanceBPNN(BPNNOverSample(ip))
    # SMOTE
    ip.performanceCNN(CNNSMOTE(ip))
    ip.performanceAE(AESMOTE(ip))
    ip.performanceDAE(DAESMOTE(ip))
    ip.performanceRNN(RNNSMOTE(ip))
    ip.performanceLSTM(LSTMSMOTE(ip))
    ip.performanceBPNN(BPNNSMOTE(ip))
    # ADASYN
    ip.performanceCNN(CNNADASYN(ip))
    ip.performanceAE(AEADASYN(ip))
    ip.performanceDAE(DAEADASYN(ip))
    ip.performanceRNN(RNNADASYN(ip))
    ip.performanceLSTM(LSTMADASYN(ip))
    ip.performanceBPNN(BPNNADASYN(ip))
    comparision = ip.show_comparison_eposh()

    names_1 = ["CNN", "AE", "DAE", "RNN", "LSTM", "BPNN"]
    names_2 = ["Imbalanced", "UnderSampler", "OverSampler", "SMOTE", "ADASYN"]

    # Strategies by models and epochs
    # Imbalanced
    # epochs 1
    title = "Scores by metrics in DL models Imbalanced with 1 epoch"
    pos = [0, 1, 2, 3, 4, 5]
    path = path_save + "Experimento4/Imbalanced/DL_imbalanced_1_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 20
    title = "Scores by metrics in DL models Imbalanced with 20 epochs"
    pos = [30, 31, 32, 33, 34, 35]
    path = path_save + "Experimento4/Imbalanced/DL_imbalanced_20_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 50
    title = "Scores by metrics in DL models Imbalanced with 50 epochs"
    pos = [60, 61, 62, 63, 64, 65]
    path = path_save + "Experimento4/Imbalanced/DL_imbalanced_50_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 100
    title = "Scores by metrics in DL models Imbalanced with 100 epochs"
    pos = [90, 91, 92, 93, 94, 95]
    path = path_save + "Experimento4/Imbalanced/DL_imbalanced_100_test"
    plot_bar_test(ip, names_1, title, pos, path)

    # UnderSampler
    # epoch 1
    title = "Scores by metrics in DL models UnderSampler with 1 epoch"
    pos = [6, 7, 8, 9, 10, 11]
    path = path_save + "Experimento4/UnderSampler/DL_UnderSampler_1_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 20
    title = "Scores by metrics in DL models UnderSampler with 20 epochs"
    pos = [36, 37, 38, 39, 40, 41]
    path = path_save + "Experimento4/UnderSampler/DL_UnderSampler_20_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 50
    title = "Scores by metrics in DL models UnderSampler with 50 epochs"
    pos = [66, 67, 68, 69, 70, 71]
    path = path_save + "Experimento4/UnderSampler/DL_UnderSampler_50_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 100
    title = "Scores by metrics in DL models UnderSampler with 100 epochs"
    pos = [96, 97, 98, 99, 100, 101]
    path = path_save + "Experimento4/UnderSampler/DL_UnderSampler_100_test"
    plot_bar_test(ip, names_1, title, pos, path)

    # OverSampler
    # epoch 1
    title = "Scores by metrics in DL models OverSampler with 1 epoch"
    pos = [12, 13, 14, 15, 16, 17]
    path = path_save + "Experimento4/OverSampler/DL_OverSampler_1_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 20
    title = "Scores by metrics in DL models OverSampler with 20 epochs"
    pos = [42, 43, 44, 45, 46, 47]
    path = path_save + "Experimento4/OverSampler/DL_OverSampler_20_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 50
    title = "Scores by metrics in DL models OverSampler with 50 epochs"
    pos = [72, 73, 74, 75, 76, 77]
    path = path_save + "Experimento4/OverSampler/DL_OverSampler_50_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 100
    title = "Scores by metrics in DL models OverSampler with 100 epochs"
    pos = [102, 103, 104, 105, 106, 107]
    path = path_save + "Experimento4/OverSampler/DL_OverSampler_100_test"
    plot_bar_test(ip, names_1, title, pos, path)

    # SMOTE
    # epoch 1
    title = "Scores by metrics in DL models SMOTE with 1 epoch"
    pos = [18, 19, 20, 21, 22, 23]
    path = path_save + "Experimento4/SMOTE/DL_SMOTE_1_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 20
    title = "Scores by metrics in DL models SMOTE with 20 epochs"
    pos = [48, 49, 50, 51, 52, 53]
    path = path_save + "Experimento4/SMOTE/DL_SMOTE_20_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 50
    title = "Scores by metrics in DL models SMOTE with 50 epochs"
    pos = [78, 79, 80, 81, 82, 83]
    path = path_save + "Experimento4/SMOTE/DL_SMOTE_50_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 100
    title = "Scores by metrics in DL models SMOTE with 100 epochs"
    pos = [108, 109, 110, 111, 112, 113]
    path = path_save + "Experimento4/SMOTE/DL_SMOTE_100_test"
    plot_bar_test(ip, names_1, title, pos, path)

    # ADASYN
    # epoch 1
    title = "Scores by metrics in DL models ADASYN with 1 epoch"
    pos = [24, 25, 26, 27, 28, 29]
    path = path_save + "Experimento4/ADASYN/DL_ADASYN_1_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 20
    title = "Scores by metrics in DL models ADASYN with 20 epochs"
    pos = [54, 55, 56, 57, 58, 59]
    path = path_save + "Experimento4/ADASYN/DL_ADASYN_20_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 50
    title = "Scores by metrics in DL models ADASYN with 50 epochs"
    pos = [84, 85, 86, 87, 88, 89]
    path = path_save + "Experimento4/ADASYN/DL_ADASYN_50_test"
    plot_bar_test(ip, names_1, title, pos, path)
    # epochs 100
    title = "Scores by metrics in DL models ADASYN with 100 epochs"
    pos = [114, 115, 116, 117, 118, 119]
    path = path_save + "Experimento4/ADASYN/DL_ADASYN_100_test"
    plot_bar_test(ip, names_1, title, pos, path)

    # every model by epochs and strategies
    # CNN
    # epoch 1
    title = "Scores by metrics in CNN models with 1 epoch"
    pos = [0, 6, 12, 16, 24]
    path = path_save + "Experimento4/CNN/CNN_1_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 20
    title = "Scores by metrics in CNN models with 20 epochs"
    pos = [30, 36, 42, 48, 54]
    path = path_save + "Experimento4/CNN/CNN_20_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 50
    title = "Scores by metrics in CNN models with 50 epochs"
    pos = [60, 66, 72, 78, 84]
    path = path_save + "Experimento4/CNN/CNN_50_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 100
    title = "Scores by metrics in CNN models with 100 epochs"
    pos = [90, 96, 102, 108, 114]
    path = path_save + "Experimento4/CNN/CNN_100_test"
    plot_bar_test(ip, names_2, title, pos, path)

    # AE
    # epoch 1
    title = "Scores by metrics in AE models with 1 epoch"
    pos = [1, 7, 13, 17, 25]
    path = path_save + "Experimento4/AE/AE_1_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 20
    title = "Scores by metrics in AE models with 20 epochs"
    pos = [31, 37, 43, 49, 55]
    path = path_save + "Experimento4/AE/AE_20_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 50
    title = "Scores by metrics in AE models with 50 epochs"
    pos = [61, 67, 73, 79, 85]
    path = path_save + "Experimento4/AE/AE_50_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 100
    title = "Scores by metrics in AE models with 100 epochs"
    pos = [91, 97, 103, 109, 115]
    path = path_save + "Experimento4/AE/AE_100_test"
    plot_bar_test(ip, names_2, title, pos, path)

    # DAE
    # epoch 1
    title = "Scores by metrics in DAE models with 1 epoch"
    pos = [2, 8, 14, 20, 26]
    path = path_save + "Experimento4/DAE/DAE_1_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 20
    title = "Scores by metrics in DAE models with 20 epochs"
    pos = [32, 38, 44, 50, 56]
    path = path_save + "Experimento4/DAE/DAE_20_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 50
    title = "Scores by metrics in DAE models with 50 epochs"
    pos = [62, 68, 74, 80, 86]
    path = path_save + "Experimento4/DAE/DAE_50_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 100
    title = "Scores by metrics in DAE models with 100 epochs"
    pos = [92, 98, 104, 110, 116]
    path = path_save + "Experimento4/DAE/DAE_100_test"
    plot_bar_test(ip, names_2, title, pos, path)

    # RNN
    # epoch 1
    title = "Scores by metrics in RNN models with 1 epoch"
    pos = [3, 9, 15, 21, 27]
    path = path_save + "Experimento4/RNN/RNN_1_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 20
    title = "Scores by metrics in RNN models with 20 epochs"
    pos = [33, 39, 45, 51, 57]
    path = path_save + "Experimento4/RNN/RNN_20_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 50
    title = "Scores by metrics in RNN models with 50 epochs"
    pos = [63, 69, 75, 81, 87]
    path = path_save + "Experimento4/RNN/RNN_50_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 100
    title = "Scores by metrics in RNN models with 100 epochs"
    pos = [93, 99, 105, 111, 117]
    path = path_save + "Experimento4/RNN/RNN_100_test"
    plot_bar_test(ip, names_2, title, pos, path)

    # LSTM
    # epoch 1
    title = "Scores by metrics in LSTM models with 1 epoch"
    pos = [4, 10, 16, 22, 28]
    path = path_save + "Experimento4/LSTM/LSTM_1_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 20
    title = "Scores by metrics in LSTM models with 20 epochs"
    pos = [34, 40, 46, 52, 58]
    path = path_save + "Experimento4/LSTM/LSTM_20_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 50
    title = "Scores by metrics in LSTM models with 50 epochs"
    pos = [64, 70, 76, 82, 88]
    path = path_save + "Experimento4/LSTM/LSTM_50_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 100
    title = "Scores by metrics in LSTM models with 100 epochs"
    pos = [94, 100, 106, 112, 118]
    path = path_save + "Experimento4/LSTM/LSTM_100_test"
    plot_bar_test(ip, names_2, title, pos, path)

    # BPNN
    # epoch 1
    title = "Scores by metrics in BPNN models with 1 epoch"
    pos = [5, 11, 17, 23, 29]
    path = path_save + "Experimento4/BPNN/BPNN_1_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 20
    title = "Scores by metrics in BPNN models with 20 epochs"
    pos = [35, 41, 47, 53, 59]
    path = path_save + "Experimento4/BPNN/BPNN_20_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 50
    title = "Scores by metrics in BPNN models with 50 epochs"
    pos = [65, 71, 77, 83, 89]
    path = path_save + "Experimento4/BPNN/BPNN_50_test"
    plot_bar_test(ip, names_2, title, pos, path)
    # epochs 100
    title = "Scores by metrics in BPNN models with 100 epochs"
    pos = [95, 101, 107, 113, 119]
    path = path_save + "Experimento4/BPNN/BPNN_100_test"
    plot_bar_test(ip, names_2, title, pos, path)

    # every models by strategies
    epochs = [1, 20, 50, 100]
    # CNN
    # Imbalanced
    title = "Scores by metrics in CNN models with Imbalanced"
    pos = [0, 30, 60, 90]
    path = path_save + "Experimento4/CNN/CNN_Imbalanced_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # UnderSampler
    title = "Scores by metrics in CNN models with UnderSampler"
    pos = [6, 36, 66, 96]
    path = path_save + "Experimento4/CNN/CNN_UnderSampler_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # OverSampler
    title = "Scores by metrics in CNN models with OverSampler"
    pos = [12, 42, 72, 102]
    path = path_save + "Experimento4/CNN/CNN_OverSampler_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # SMOTE
    title = "Scores by metrics in CNN models with SMOTE"
    pos = [18, 48, 78, 108]
    path = path_save + "Experimento4/CNN/CNN_SMOTE_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # ADASYN
    title = "Scores by metrics in CNN models with ADASYN"
    pos = [24, 54, 84, 114]
    path = path_save + "Experimento4/CNN/CNN_ADASYN_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # AE
    # Imbalanced
    title = "Scores by metrics in AE models with Imbalanced"
    pos = [1, 31, 61, 91]
    path = path_save + "Experimento4/AE/AE_Imbalanced_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # UnderSampler
    title = "Scores by metrics in AE models with UnderSampler"
    pos = [7, 37, 67, 97]
    path = path_save + "Experimento4/AE/AE_UnderSampler_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # OverSampler
    title = "Scores by metrics in AE models with OverSampler"
    pos = [13, 43, 73, 103]
    path = path_save + "Experimento4/AE/AE_OverSampler_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # SMOTE
    title = "Scores by metrics in AE models with SMOTE"
    pos = [19, 49, 79, 109]
    path = path_save + "Experimento4/AE/AE_SMOTE_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # ADASYN
    title = "Scores by metrics in AE models with ADASYN"
    pos = [25, 55, 85, 115]
    path = path_save + "Experimento4/AE/AE_ADASYN_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # DAE
    # Imbalanced
    title = "Scores by metrics in DAE models with Imbalanced"
    pos = [2, 32, 62, 92]
    path = path_save + "Experimento4/DAE/DAE_Imbalanced_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # UnderSampler
    title = "Scores by metrics in DAE models with UnderSampler"
    pos = [8, 38, 68, 98]
    path = path_save + "Experimento4/DAE/DAE_UnderSampler_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # OverSampler
    title = "Scores by metrics in DAE models with OverSampler"
    pos = [14, 44, 74, 104]
    path = path_save + "Experimento4/DAE/DAE_OverSampler_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # SMOTE
    title = "Scores by metrics in DAE models with SMOTE"
    pos = [20, 50, 80, 110]
    path = path_save + "Experimento4/DAE/DAE_SMOTE_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # ADASYN
    title = "Scores by metrics in DAE models with ADASYN"
    pos = [26, 56, 86, 116]
    path = path_save + "Experimento4/DAE/DAE_ADASYN_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # RNN
    # Imbalanced
    title = "Scores by metrics in RNN models with Imbalanced"
    pos = [3, 33, 63, 93]
    path = path_save + "Experimento4/RNN/RNN_Imbalanced_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # UnderSampler
    title = "Scores by metrics in RNN models with UnderSampler"
    pos = [9, 39, 69, 99]
    path = path_save + "Experimento4/RNN/RNN_UnderSampler_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # OverSampler
    title = "Scores by metrics in RNN models with OverSampler"
    pos = [15, 45, 75, 105]
    path = path_save + "Experimento4/RNN/RNN_OverSampler_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # SMOTE
    title = "Scores by metrics in RNN models with SMOTE"
    pos = [21, 51, 81, 111]
    path = path_save + "Experimento4/RNN/RNN_SMOTE_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # ADASYN
    title = "Scores by metrics in RNN models with ADASYN"
    pos = [27, 57, 87, 117]
    path = path_save + "Experimento4/RNN/RNN_ADASYN_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # LSTM
    # Imbalanced
    title = "Scores by metrics in LSTM models with Imbalanced"
    pos = [4, 34, 64, 94]
    path = path_save + "Experimento4/LSTM/LSTM_Imbalanced_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # UnderSampler
    title = "Scores by metrics in LSTM models with UnderSampler"
    pos = [10, 40, 70, 100]
    path = path_save + "Experimento4/LSTM/LSTM_UnderSampler_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # OverSampler
    title = "Scores by metrics in LSTM models with OverSampler"
    pos = [16, 46, 76, 106]
    path = path_save + "Experimento4/LSTM/LSTM_OverSampler_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # SMOTE
    title = "Scores by metrics in LSTM models with SMOTE"
    pos = [22, 52, 82, 112]
    path = path_save + "Experimento4/LSTM/LSTM_SMOTE_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # ADASYN
    title = "Scores by metrics in LSTM models with ADASYN"
    pos = [28, 58, 88, 118]
    path = path_save + "Experimento4/LSTM/LSTM_ADASYN_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # BPNN
    # Imbalanced
    title = "Scores by metrics in BPNN models with Imbalanced"
    pos = [5, 35, 65, 95]
    path = path_save + "Experimento4/BPNN/BPNN_Imbalanced_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # UnderSampler
    title = "Scores by metrics in BPNN models with UnderSampler"
    pos = [11, 41, 71, 101]
    path = path_save + "Experimento4/BPNN/BPNN_UnderSampler_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # OverSampler
    title = "Scores by metrics in BPNN models with OverSampler"
    pos = [17, 47, 77, 107]
    path = path_save + "Experimento4/BPNN/BPNN_OverSampler_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # SMOTE
    title = "Scores by metrics in BPNN models with SMOTE"
    pos = [23, 53, 83, 113]
    path = path_save + "Experimento4/BPNN/BPNN_SMOTE_test"
    plot_bar_test(ip, epochs, title, pos, path)
    # ADASYN
    title = "Scores by metrics in BPNN models with ADASYN"
    pos = [29, 59, 89, 119]
    path = path_save + "Experimento4/BPNN/BPNN_ADASYN_test"
    plot_bar_test(ip, epochs, title, pos, path)

    values = comparision.values
    df1_1 = pd.DataFrame(values[:6], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                              'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df1_2 = pd.DataFrame(values[6:12], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df1_3 = pd.DataFrame(values[12:18], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df1_4 = pd.DataFrame(values[18:24], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)

    df2_1 = pd.DataFrame(values[24:30], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df2_2 = pd.DataFrame(values[30:36], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df2_3 = pd.DataFrame(values[36:42], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df2_4 = pd.DataFrame(values[42:48], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)

    df3_1 = pd.DataFrame(values[48:54], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df3_2 = pd.DataFrame(values[54:60], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df3_3 = pd.DataFrame(values[60:66], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df3_4 = pd.DataFrame(values[66:72], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)

    df4_1 = pd.DataFrame(values[72:78], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df4_2 = pd.DataFrame(values[78:84], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df4_3 = pd.DataFrame(values[84:90], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df4_4 = pd.DataFrame(values[90:96], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                 'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)

    df5_1 = pd.DataFrame(values[96:102], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                  'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df5_2 = pd.DataFrame(values[102:108], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                   'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df5_3 = pd.DataFrame(values[108:114], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                   'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)
    df5_4 = pd.DataFrame(values[114:120], columns=['Eposh', 'Model', 'Accuracy', 'AUC', 'Precision Score',
                                                   'Recall Score', 'F1 Score']).sort_values('F1 Score', ascending=False)

    book = load_workbook('Experimentos.xlsx')
    writer = pd.ExcelWriter('Experimentos.xlsx', engine='openpyxl', if_sheet_exists='replace')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df1_1.to_excel(writer, sheet_name='Experimento4', index=False, startrow=0)
    df1_2.to_excel(writer, sheet_name='Experimento4', index=False, startrow=9)
    df1_3.to_excel(writer, sheet_name='Experimento4', index=False, startrow=18)
    df1_4.to_excel(writer, sheet_name='Experimento4', index=False, startrow=27)
    df2_1.to_excel(writer, sheet_name='Experimento4', index=False, startrow=36)
    df2_2.to_excel(writer, sheet_name='Experimento4', index=False, startrow=45)
    df2_3.to_excel(writer, sheet_name='Experimento4', index=False, startrow=54)
    df2_4.to_excel(writer, sheet_name='Experimento4', index=False, startrow=63)
    df3_1.to_excel(writer, sheet_name='Experimento4', index=False, startrow=72)
    df3_2.to_excel(writer, sheet_name='Experimento4', index=False, startrow=81)
    df3_3.to_excel(writer, sheet_name='Experimento4', index=False, startrow=90)
    df3_4.to_excel(writer, sheet_name='Experimento4', index=False, startrow=99)
    df4_1.to_excel(writer, sheet_name='Experimento4', index=False, startrow=108)
    df4_2.to_excel(writer, sheet_name='Experimento4', index=False, startrow=117)
    df4_3.to_excel(writer, sheet_name='Experimento4', index=False, startrow=126)
    df4_4.to_excel(writer, sheet_name='Experimento4', index=False, startrow=135)
    df5_1.to_excel(writer, sheet_name='Experimento4', index=False, startrow=144)
    df5_2.to_excel(writer, sheet_name='Experimento4', index=False, startrow=153)
    df5_3.to_excel(writer, sheet_name='Experimento4', index=False, startrow=162)
    df5_4.to_excel(writer, sheet_name='Experimento4', index=False, startrow=171)
    comparision = comparision.sort_values('F1 Score', ascending=False)
    comparision.to_excel(writer, sheet_name='Experimento4', index=False, startcol=8)
    writer.save()
    writer.close()

    comparision = ip.show_comparison_train()

    # Strategies by models and epochs
    # Imbalanced
    # epochs 1
    title = "Scores by metrics in DL models Imbalanced with 1 epoch"
    pos = [0, 1, 2, 3, 4, 5]
    path = path_save + "Experimento5/Imbalanced/DL_imbalanced_1_train"
    plot_bar_train(ip, names_1, title, pos, path)
    # epochs 20
    title = "Scores by metrics in DL models Imbalanced with 20 epochs"
    pos = [30, 31, 32, 33, 34, 35]
    path = path_save + "Experimento5/Imbalanced/DL_imbalanced_20_train"
    plot_graph(ip, names_1, title, pos, 20, path)
    # epochs 50
    title = "Scores by metrics in DL models Imbalanced with 50 epochs"
    pos = [60, 61, 62, 63, 64, 65]
    path = path_save + "Experimento5/Imbalanced/DL_imbalanced_50_train"
    plot_graph(ip, names_1, title, pos, 50, path)
    # epochs 100
    title = "Scores by metrics in DL models Imbalanced with 100 epochs"
    pos = [90, 91, 92, 93, 94, 95]
    path = path_save + "Experimento5/Imbalanced/DL_imbalanced_100_train"
    plot_graph(ip, names_1, title, pos, 100, path)
    # UnderSampler
    # epoch 1
    title = "Scores by metrics in DL models UnderSampler with 1 epoch"
    pos = [6, 7, 8, 9, 10, 11]
    path = path_save + "Experimento5/UnderSampler/DL_UnderSampler_1_train"
    plot_bar_train(ip, names_1, title, pos, path)
    # epochs 20
    title = "Scores by metrics in DL models UnderSampler with 20 epochs"
    pos = [36, 37, 38, 39, 40, 41]
    path = path_save + "Experimento5/UnderSampler/DL_UnderSampler_20_train"
    plot_graph(ip, names_1, title, pos, 20, path)
    # epochs 50
    title = "Scores by metrics in DL models UnderSampler with 50 epochs"
    pos = [66, 67, 68, 69, 70, 71]
    path = path_save + "Experimento5/UnderSampler/DL_UnderSampler_50_train"
    plot_graph(ip, names_1, title, pos, 50, path)
    # epochs 100
    title = "Scores by metrics in DL models UnderSampler with 100 epochs"
    pos = [96, 97, 98, 99, 100, 101]
    path = path_save + "Experimento5/UnderSampler/DL_UnderSampler_100_train"
    plot_graph(ip, names_1, title, pos, 100, path)
    # OverSampler
    # epoch 1
    title = "Scores by metrics in DL models OverSampler with 1 epoch"
    pos = [12, 13, 14, 15, 16, 17]
    path = path_save + "Experimento5/OverSampler/DL_OverSampler_1_train"
    plot_bar_train(ip, names_1, title, pos, path)
    # epochs 20
    title = "Scores by metrics in DL models OverSampler with 20 epochs"
    pos = [42, 43, 44, 45, 46, 47]
    path = path_save + "Experimento5/OverSampler/DL_OverSampler_20_train"
    plot_graph(ip, names_1, title, pos, 20, path)
    # epochs 50
    title = "Scores by metrics in DL models OverSampler with 50 epochs"
    pos = [72, 73, 74, 75, 76, 77]
    path = path_save + "Experimento5/OverSampler/DL_OverSampler_50_train"
    plot_graph(ip, names_1, title, pos, 50, path)
    # epochs 100
    title = "Scores by metrics in DL models OverSampler with 100 epochs"
    pos = [102, 103, 104, 105, 106, 107]
    path = path_save + "Experimento5/OverSampler/DL_OverSampler_100_train"
    plot_graph(ip, names_1, title, pos, 100, path)
    # SMOTE
    # epoch 1
    title = "Scores by metrics in DL models SMOTE with 1 epoch"
    pos = [18, 19, 20, 21, 22, 23]
    path = path_save + "Experimento5/SMOTE/DL_SMOTE_1_train"
    plot_bar_train(ip, names_1, title, pos, path)
    # epochs 20
    title = "Scores by metrics in DL models SMOTE with 20 epochs"
    pos = [48, 49, 50, 51, 52, 53]
    path = path_save + "Experimento5/SMOTE/DL_SMOTE_20_train"
    plot_graph(ip, names_1, title, pos, 20, path)
    # epochs 50
    title = "Scores by metrics in DL models SMOTE with 50 epochs"
    pos = [78, 79, 80, 81, 82, 83]
    path = path_save + "Experimento5/SMOTE/DL_SMOTE_50_train"
    plot_graph(ip, names_1, title, pos, 50, path)
    # epochs 100
    title = "Scores by metrics in DL models SMOTE with 100 epochs"
    pos = [108, 109, 110, 111, 112, 113]
    path = path_save + "Experimento5/SMOTE/DL_SMOTE_100_train"
    plot_graph(ip, names_1, title, pos, 100, path)
    # ADASYN
    # epoch 1
    title = "Scores by metrics in DL models ADASYN with 1 epoch"
    pos = [24, 25, 26, 27, 28, 29]
    path = path_save + "Experimento5/ADASYN/DL_ADASYN_1_train"
    plot_bar_train(ip, names_1, title, pos, path)
    # epochs 20
    title = "Scores by metrics in DL models ADASYN with 20 epochs"
    pos = [54, 55, 56, 57, 58, 59]
    path = path_save + "Experimento5/ADASYN/DL_ADASYN_20_train"
    plot_graph(ip, names_1, title, pos, 20, path)
    # epochs 50
    title = "Scores by metrics in DL models ADASYN with 50 epochs"
    pos = [84, 85, 86, 87, 88, 89]
    path = path_save + "Experimento5/ADASYN/DL_ADASYN_50_train"
    plot_graph(ip, names_1, title, pos, 50, path)
    # epochs 100
    title = "Scores by metrics in DL models ADASYN with 100 epochs"
    pos = [114, 115, 116, 117, 118, 119]
    path = path_save + "Experimento5/ADASYN/DL_ADASYN_100_train"
    plot_graph(ip, names_1, title, pos, 100, path)

    # every model by epochs and strategies
    # CNN
    # epoch 1
    title = "Scores by metrics in CNN models with 1 epoch"
    pos = [0, 6, 12, 16, 24]
    path = path_save + "Experimento5/CNN/CNN_1_train"
    plot_bar_train(ip, names_2, title, pos, path)
    # epochs 20
    title = "Scores by metrics in CNN models with 20 epochs"
    pos = [30, 36, 42, 48, 54]
    path = path_save + "Experimento5/CNN/CNN_20_train"
    plot_graph(ip, names_2, title, pos, 20, path)
    # epochs 50
    title = "Scores by metrics in CNN models with 50 epochs"
    pos = [60, 66, 72, 78, 84]
    path = path_save + "Experimento5/CNN/CNN_50_train"
    plot_graph(ip, names_2, title, pos, 50, path)
    # epochs 100
    title = "Scores by metrics in CNN models with 100 epochs"
    pos = [90, 96, 102, 108, 114]
    path = path_save + "Experimento5/CNN/CNN_100_train"
    plot_graph(ip, names_2, title, pos, 100, path)
    # AE
    # epoch 1
    title = "Scores by metrics in AE models with 1 epoch"
    pos = [1, 7, 13, 17, 25]
    path = path_save + "Experimento5/AE/AE_1_train"
    plot_bar_train(ip, names_2, title, pos, path)
    # epochs 20
    title = "Scores by metrics in AE models with 20 epochs"
    pos = [31, 37, 43, 49, 55]
    path = path_save + "Experimento5/AE/AE_20_train"
    plot_graph(ip, names_2, title, pos, 20, path)
    # epochs 50
    title = "Scores by metrics in AE models with 50 epochs"
    pos = [61, 67, 73, 79, 85]
    path = path_save + "Experimento5/AE/AE_50_train"
    plot_graph(ip, names_2, title, pos, 50, path)
    # epochs 100
    title = "Scores by metrics in AE models with 100 epochs"
    pos = [91, 97, 103, 109, 115]
    path = path_save + "Experimento5/AE/AE_100_train"
    plot_graph(ip, names_2, title, pos, 100, path)
    # DAE
    # epoch 1
    title = "Scores by metrics in DAE models with 1 epoch"
    pos = [2, 8, 14, 20, 26]
    path = path_save + "Experimento5/DAE/DAE_1_train"
    plot_bar_train(ip, names_2, title, pos, path)
    # epochs 20
    title = "Scores by metrics in DAE models with 20 epochs"
    pos = [32, 38, 44, 50, 56]
    path = path_save + "Experimento5/DAE/DAE_20_train"
    plot_graph(ip, names_2, title, pos, 20, path)
    # epochs 50
    title = "Scores by metrics in DAE models with 50 epochs"
    pos = [62, 68, 74, 80, 86]
    path = path_save + "Experimento5/DAE/DAE_50_train"
    plot_graph(ip, names_2, title, pos, 50, path)
    # epochs 100
    title = "Scores by metrics in DAE models with 100 epochs"
    pos = [92, 98, 104, 110, 116]
    path = path_save + "Experimento5/DAE/DAE_100_train"
    plot_graph(ip, names_2, title, pos, 100, path)
    # RNN
    # epoch 1
    title = "Scores by metrics in RNN models with 1 epoch"
    pos = [3, 9, 15, 21, 27]
    path = path_save + "Experimento5/RNN/RNN_1_train"
    plot_bar_train(ip, names_2, title, pos, path)
    # epochs 20
    title = "Scores by metrics in RNN models with 20 epochs"
    pos = [33, 39, 45, 51, 57]
    path = path_save + "Experimento5/RNN/RNN_20_train"
    plot_graph(ip, names_2, title, pos, 20, path)
    # epochs 50
    title = "Scores by metrics in RNN models with 50 epochs"
    pos = [63, 69, 75, 81, 87]
    path = path_save + "Experimento5/RNN/RNN_50_train"
    plot_graph(ip, names_2, title, pos, 50, path)
    # epochs 100
    title = "Scores by metrics in RNN models with 100 epochs"
    pos = [93, 99, 105, 111, 117]
    path = path_save + "Experimento5/RNN/RNN_100_train"
    plot_graph(ip, names_2, title, pos, 100, path)
    # LSTM
    # epoch 1
    title = "Scores by metrics in LSTM models with 1 epoch"
    pos = [4, 10, 16, 22, 28]
    path = path_save + "Experimento5/LSTM/LSTM_1_train"
    plot_bar_train(ip, names_2, title, pos, path)
    # epochs 20
    title = "Scores by metrics in LSTM models with 20 epochs"
    pos = [34, 40, 46, 52, 58]
    path = path_save + "Experimento5/LSTM/LSTM_20_train"
    plot_graph(ip, names_2, title, pos, 20, path)
    # epochs 50
    title = "Scores by metrics in LSTM models with 50 epochs"
    pos = [64, 70, 76, 82, 88]
    path = path_save + "Experimento5/LSTM/LSTM_50_train"
    plot_graph(ip, names_2, title, pos, 50, path)
    # epochs 100
    title = "Scores by metrics in LSTM models with 100 epochs"
    pos = [94, 100, 106, 112, 118]
    path = path_save + "Experimento5/LSTM/LSTM_100_train"
    plot_graph(ip, names_2, title, pos, 100, path)
    # BPNN
    # epoch 1
    title = "Scores by metrics in BPNN models with 1 epoch"
    pos = [5, 11, 17, 23, 29]
    path = path_save + "Experimento5/BPNN/BPNN_1_train"
    plot_bar_train(ip, names_2, title, pos, path)
    # epochs 20
    title = "Scores by metrics in BPNN models with 20 epochs"
    pos = [35, 41, 47, 53, 59]
    path = path_save + "Experimento5/BPNN/BPNN_20_train"
    plot_graph(ip, names_2, title, pos, 20, path)
    # epochs 50
    title = "Scores by metrics in BPNN models with 50 epochs"
    pos = [65, 71, 77, 83, 89]
    path = path_save + "Experimento5/BPNN/BPNN_50_train"
    plot_graph(ip, names_2, title, pos, 50, path)
    # epochs 100
    title = "Scores by metrics in BPNN models with 100 epochs"
    pos = [95, 101, 107, 113, 119]
    path = path_save + "Experimento5/BPNN/BPNN_100_train"
    plot_graph(ip, names_2, title, pos, 100, path)

    # every models by strategies
    # CNN
    # Imbalanced
    title = "Scores by metrics in CNN models with Imbalanced"
    pos = [0, 30, 60, 90]
    path = path_save + "Experimento5/CNN/CNN_Imbalanced_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # UnderSampler
    title = "Scores by metrics in CNN models with UnderSampler"
    pos = [6, 36, 66, 96]
    path = path_save + "Experimento5/CNN/CNN_UnderSampler_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # OverSampler
    title = "Scores by metrics in CNN models with OverSampler"
    pos = [12, 42, 72, 102]
    path = path_save + "Experimento5/CNN/CNN_OverSampler_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # SMOTE
    title = "Scores by metrics in CNN models with SMOTE"
    pos = [18, 48, 78, 108]
    path = path_save + "Experimento5/CNN/CNN_SMOTE_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # ADASYN
    title = "Scores by metrics in CNN models with ADASYN"
    pos = [24, 54, 84, 114]
    path = path_save + "Experimento5/CNN/CNN_ADASYN_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # AE
    # Imbalanced
    title = "Scores by metrics in AE models with Imbalanced"
    pos = [1, 31, 61, 91]
    path = path_save + "Experimento5/AE/AE_Imbalanced_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # UnderSampler
    title = "Scores by metrics in AE models with UnderSampler"
    pos = [7, 37, 67, 97]
    path = path_save + "Experimento5/AE/AE_UnderSampler_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # OverSampler
    title = "Scores by metrics in AE models with OverSampler"
    pos = [13, 43, 73, 103]
    path = path_save + "Experimento5/AE/AE_OverSampler_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # SMOTE
    title = "Scores by metrics in AE models with SMOTE"
    pos = [19, 49, 79, 109]
    path = path_save + "Experimento5/AE/AE_SMOTE_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # ADASYN
    title = "Scores by metrics in AE models with ADASYN"
    pos = [25, 55, 85, 115]
    path = path_save + "Experimento5/AE/AE_ADASYN_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # DAE
    # Imbalanced
    title = "Scores by metrics in DAE models with Imbalanced"
    pos = [2, 32, 62, 92]
    path = path_save + "Experimento5/DAE/DAE_Imbalanced_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # UnderSampler
    title = "Scores by metrics in DAE models with UnderSampler"
    pos = [8, 38, 68, 98]
    path = path_save + "Experimento5/DAE/DAE_UnderSampler_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # OverSampler
    title = "Scores by metrics in DAE models with OverSampler"
    pos = [14, 44, 74, 104]
    path = path_save + "Experimento5/DAE/DAE_OverSampler_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # SMOTE
    title = "Scores by metrics in DAE models with SMOTE"
    pos = [20, 50, 80, 110]
    path = path_save + "Experimento5/DAE/DAE_SMOTE_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # ADASYN
    title = "Scores by metrics in DAE models with ADASYN"
    pos = [26, 56, 86, 116]
    path = path_save + "Experimento5/DAE/DAE_ADASYN_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # RNN
    # Imbalanced
    title = "Scores by metrics in RNN models with Imbalanced"
    pos = [3, 33, 63, 93]
    path = path_save + "Experimento5/RNN/RNN_Imbalanced_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # UnderSampler
    title = "Scores by metrics in RNN models with UnderSampler"
    pos = [9, 39, 69, 99]
    path = path_save + "Experimento5/RNN/RNN_UnderSampler_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # OverSampler
    title = "Scores by metrics in RNN models with OverSampler"
    pos = [15, 45, 75, 105]
    path = path_save + "Experimento5/RNN/RNN_OverSampler_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # SMOTE
    title = "Scores by metrics in RNN models with SMOTE"
    pos = [21, 51, 81, 111]
    path = path_save + "Experimento5/RNN/RNN_SMOTE_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # ADASYN
    title = "Scores by metrics in RNN models with ADASYN"
    pos = [27, 57, 87, 117]
    path = path_save + "Experimento5/RNN/RNN_ADASYN_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # LSTM
    # Imbalanced
    title = "Scores by metrics in LSTM models with Imbalanced"
    pos = [4, 34, 64, 94]
    path = path_save + "Experimento5/LSTM/LSTM_Imbalanced_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # UnderSampler
    title = "Scores by metrics in LSTM models with UnderSampler"
    pos = [10, 40, 70, 100]
    path = path_save + "Experimento5/LSTM/LSTM_UnderSampler_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # OverSampler
    title = "Scores by metrics in LSTM models with OverSampler"
    pos = [16, 46, 76, 106]
    path = path_save + "Experimento5/LSTM/LSTM_OverSampler_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # SMOTE
    title = "Scores by metrics in LSTM models with SMOTE"
    pos = [22, 52, 82, 112]
    path = path_save + "Experimento5/LSTM/LSTM_SMOTE_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # ADASYN
    title = "Scores by metrics in LSTM models with ADASYN"
    pos = [28, 58, 88, 118]
    path = path_save + "Experimento5/LSTM/LSTM_ADASYN_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # BPNN
    # Imbalanced
    title = "Scores by metrics in BPNN models with Imbalanced"
    pos = [5, 35, 65, 95]
    path = path_save + "Experimento5/BPNN/BPNN_Imbalanced_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # UnderSampler
    title = "Scores by metrics in BPNN models with UnderSampler"
    pos = [11, 41, 71, 101]
    path = path_save + "Experimento5/BPNN/BPNN_UnderSampler_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # OverSampler
    title = "Scores by metrics in BPNN models with OverSampler"
    pos = [17, 47, 77, 107]
    path = path_save + "Experimento5/BPNN/BPNN_OverSampler_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # SMOTE
    title = "Scores by metrics in BPNN models with SMOTE"
    pos = [23, 53, 83, 113]
    path = path_save + "Experimento5/BPNN/BPNN_SMOTE_train"
    plot_bar_train(ip, epochs, title, pos, path)
    # ADASYN
    title = "Scores by metrics in BPNN models with ADASYN"
    pos = [29, 59, 89, 119]
    path = path_save + "Experimento5/BPNN/BPNN_ADASYN_train"
    plot_bar_train(ip, epochs, title, pos, path)

    values = comparision.values
    df11 = pd.DataFrame(values[:6], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                             'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                  ascending=False)
    df12 = pd.DataFrame(values[6:12], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                               'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                    ascending=False)
    df13 = pd.DataFrame(values[12:18], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)
    df14 = pd.DataFrame(values[18:24], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)

    df21 = pd.DataFrame(values[24:30], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)
    df22 = pd.DataFrame(values[30:36], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)
    df23 = pd.DataFrame(values[36:42], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)
    df24 = pd.DataFrame(values[42:48], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)

    df31 = pd.DataFrame(values[48:54], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)
    df32 = pd.DataFrame(values[54:60], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)
    df33 = pd.DataFrame(values[60:66], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)
    df34 = pd.DataFrame(values[66:72], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)

    df41 = pd.DataFrame(values[72:78], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)
    df42 = pd.DataFrame(values[78:84], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)
    df43 = pd.DataFrame(values[84:90], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)
    df44 = pd.DataFrame(values[90:96], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                     ascending=False)

    df51 = pd.DataFrame(values[96:102], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                 'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                      ascending=False)
    df52 = pd.DataFrame(values[102:108], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                  'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                       ascending=False)
    df53 = pd.DataFrame(values[108:114], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                  'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                       ascending=False)
    df54 = pd.DataFrame(values[114:120], columns=['Eposh', 'Model', 'Accuracy Train', 'Precision Score Train',
                                                  'Recall Score Train', 'F1 Score Train']).sort_values('F1 Score Train',
                                                                                                       ascending=False)
    book = load_workbook('Experimentos.xlsx')
    writer = pd.ExcelWriter('Experimentos.xlsx', engine='openpyxl', if_sheet_exists='replace')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df11.to_excel(writer, sheet_name='Experimento5', index=False, startrow=0)
    df12.to_excel(writer, sheet_name='Experimento5', index=False, startrow=8)
    df13.to_excel(writer, sheet_name='Experimento5', index=False, startrow=16)
    df14.to_excel(writer, sheet_name='Experimento5', index=False, startrow=24)
    df21.to_excel(writer, sheet_name='Experimento5', index=False, startrow=32)
    df22.to_excel(writer, sheet_name='Experimento5', index=False, startrow=40)
    df23.to_excel(writer, sheet_name='Experimento5', index=False, startrow=48)
    df24.to_excel(writer, sheet_name='Experimento5', index=False, startrow=56)
    df31.to_excel(writer, sheet_name='Experimento5', index=False, startrow=64)
    df32.to_excel(writer, sheet_name='Experimento5', index=False, startrow=72)
    df33.to_excel(writer, sheet_name='Experimento5', index=False, startrow=80)
    df34.to_excel(writer, sheet_name='Experimento5', index=False, startrow=88)
    df41.to_excel(writer, sheet_name='Experimento5', index=False, startrow=96)
    df42.to_excel(writer, sheet_name='Experimento5', index=False, startrow=104)
    df43.to_excel(writer, sheet_name='Experimento5', index=False, startrow=112)
    df44.to_excel(writer, sheet_name='Experimento5', index=False, startrow=120)
    df51.to_excel(writer, sheet_name='Experimento5', index=False, startrow=128)
    df52.to_excel(writer, sheet_name='Experimento5', index=False, startrow=136)
    df53.to_excel(writer, sheet_name='Experimento5', index=False, startrow=144)
    df54.to_excel(writer, sheet_name='Experimento5', index=False, startrow=152)
    comparision = comparision.sort_values('F1 Score Train', ascending=False)
    comparision.to_excel(writer, sheet_name='Experimento5', index=False, startcol=8)
    writer.save()
    writer.close()


def experimento6():
    ip = ImbalancedPerformanceClass()
    ip.solve_imbalanced("creditcard.csv")
    ip.performanceML(RFImbalanced(ip))
    ip.performanceML(RFOverSample(ip))
    ip.performanceML(RFSMOTE(ip))
    ip.performanceML(RFADASYN(ip))
    ip.performanceML(GNBUnderSample(ip))
    ip.epochs = 50
    ip.performanceCNN(CNNImbalanced(ip))
    ip.epochs = 100
    ip.performanceBPNN(BPNNUnderSample(ip))
    ip.performanceBPNN(BPNNOverSample(ip))
    ip.performanceBPNN(BPNNSMOTE(ip))
    ip.performanceBPNN(BPNNADASYN(ip))
    comparision = ip.show_comparison_eposh()
    comparision = comparision.sort_values('F1 Score', ascending=False)
    book = load_workbook('Experimentos.xlsx')
    writer = pd.ExcelWriter('Experimentos.xlsx', engine='openpyxl', if_sheet_exists='replace')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    comparision.to_excel(writer, sheet_name='Experimento6', index=False)
    writer.save()
    writer.close()


def experimento7():
    ip = ImbalancedPerformanceClass()
    ip.solve_imbalanced("creditcard.csv")
    ip.performanceML(DT(ip))
    ip.performanceML(GNB(ip))
    ip.performanceML(KNN(ip))
    ip.performanceML(LR(ip))
    ip.performanceML(MLP(ip))
    ip.performanceML(RF(ip))
    ip.performanceML(XGB(ip))
    ip.epochs = 1
    ip.performanceCNN(CNN(ip))
    ip.performanceAE(AE(ip))
    ip.performanceDAE(DAE(ip))
    ip.performanceRNN(RNN(ip))
    ip.performanceLSTM(LSTM(ip))
    ip.performanceBPNN(BPNN(ip))
    ip.epochs = 20
    ip.performanceCNN(CNN(ip))
    ip.performanceAE(AE(ip))
    ip.performanceDAE(DAE(ip))
    ip.performanceRNN(RNN(ip))
    ip.performanceLSTM(LSTM(ip))
    ip.performanceBPNN(BPNN(ip))
    ip.epochs = 50
    ip.performanceCNN(CNN(ip))
    ip.performanceAE(AE(ip))
    ip.performanceDAE(DAE(ip))
    ip.performanceRNN(RNN(ip))
    ip.performanceLSTM(LSTM(ip))
    ip.performanceBPNN(BPNN(ip))
    ip.epochs = 100
    ip.performanceCNN(CNN(ip))
    ip.performanceAE(AE(ip))
    ip.performanceDAE(DAE(ip))
    ip.performanceRNN(RNN(ip))
    ip.performanceLSTM(LSTM(ip))
    ip.performanceBPNN(BPNN(ip))
    comparision = ip.show_comparison_general()
    comparision = comparision.sort_values('F1 Score', ascending=False)
    book = load_workbook('Experimentos.xlsx')
    writer = pd.ExcelWriter('Experimentos.xlsx', engine='openpyxl', if_sheet_exists='replace')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    comparision.to_excel(writer, sheet_name='Experimento7', index=False)
    writer.save()
    writer.close()
